import datetime
from io import BytesIO
from pathlib import Path
import pytest
import json
from openpyxl import load_workbook
from pandas.core.common import not_none

from backend.models import TruckSheet, OrderSheet


def get_file_path(file):
    return Path(__file__).parent / 'data' / file


@pytest.fixture(autouse=True)
def setup_db(setup_users, login_admin, create_db_without_users, client):
    """
    Setup the database for the current test.
    """

    for i in range(2):
        data = dict(
            file_1=(open(get_file_path('truck_availability_test_small.xlsx'),
                         'rb'), 'sheet.xlsx')
        )

        client.post('/api/sheets/', content_type='multipart/form-data',
                    data=data)

        data2 = dict(
            file_1=(open(get_file_path('order_sheet_test_small.xlsx'), 'rb'),
                    'sheet.xlsx')
        )

        client.post('/api/sheets/', content_type='multipart/form-data',
                    data=data2)

    # The above two files are uploaded at pretty much the same time, so it is
    # not defined which one is the latest
    OrderSheet.query.get(1).upload_date += datetime.timedelta(hours=1)
    TruckSheet.query.get(1).upload_date += datetime.timedelta(hours=1)

    order1 = dict(
        truck_s_number=2, departure_time='10:00'
    )
    client.patch(f'/api/orders/{1}',
                 data=json.dumps(order1),
                 content_type='application/json')

    order2 = dict(
        truck_s_number=5, departure_time='5:00'
    )
    client.patch(f'/api/orders/{2}',
                 data=json.dumps(order2),
                 content_type='application/json')

    order3 = dict(
        truck_s_number=2, departure_time='5:00'
    )
    client.patch(f'api/orders/{3}',
                 data=json.dumps(order3),
                 content_type='application/json')

    order4 = dict(
        truck_s_number=1, departure_time='6:00'
    )
    client.patch(f'api/orders/{4}',
                 data=json.dumps(order4),
                 content_type='application/json')


def get_first_rides(client, sheet_id_or_latest):
    return client.get(f'api/reports/firstrides/{sheet_id_or_latest}')


# TODO: not implemented yet
def get_driver_wise(client, sheet_id_or_latest):
    return None


# TODO: not implemented yet
def get_full_assignment(client, sheet_id_or_latest):
    return None


def test_get_first_rides_latest(client, db):
    rv = get_first_rides(client, 'latest')
    assert rv.status_code == 200
    wb = load_workbook(filename=BytesIO(rv.data))
    ws = wb[wb.sheetnames[0]]
    assert ws['C4'].value == 'Sno'
    assert ws['F6'].value == 'ITV'
    assert ws['G6'].value is None


def test_get_first_rides_one(client, db):
    rv = get_first_rides(client, 1)
    assert rv.status_code == 200
    wb = load_workbook(filename=BytesIO(rv.data))
    ws = wb[wb.sheetnames[0]]
    assert ws['I4'].value == 'Delivery Deadline'
    assert ws['L5'].value == 'Helmond'
    assert ws['O5'].value == 'Has a personal appointment at 15.30'


def test_get_first_rides_two(client, db):
    rv = get_first_rides(client, 2)
    assert rv.status_code == 200
    wb = load_workbook(filename=BytesIO(rv.data))
    ws = wb[wb.sheetnames[0]]
    assert ws['G4'].value == 'chassis'
    assert ws['G5'].value is None
    assert ws['D5'].value is None


def test_get_first_rides_oob(client, db):
    rv = get_first_rides(client, 5)
    assert rv.status_code == 404


def test_get_first_rides_not_latest(client, db):
    rv = get_first_rides(client, 'random')
    assert rv.status_code == 404
