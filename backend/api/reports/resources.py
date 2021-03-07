import time
import io
from flask import send_file
from flask.views import MethodView
from flask_login import current_user
from sqlalchemy import func, and_
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from backend.plugins import db
from backend.models import Order, OrderSheet
from backend.extensions import Blueprint, roles_required

bp = Blueprint('reports',
               'reports',
               description='Get report like first rides assignment')


@bp.route('/firstrides/<sheet_id_or_latest>')
class FirstRides(MethodView):

    @bp.response(code=200)
    @bp.alt_response("NOT_FOUND", code=404)
    @roles_required('view-only', 'planner', 'administrator')
    def get(self, sheet_id_or_latest):
        """
        Get a workbook containing a first rides report from an order sheet.

        In case 'sheet_id_or_latest is 'latest', the most recently uploaded
        order sheet will be used.

        Required roles: view-only, planner, administrator
        """
        # Request the order sheet for this planning
        is_view_only = current_user.role == 'view-only'
        order_sheet = OrderSheet.query.get_sheet_or_404(sheet_id_or_latest,
                                                        is_view_only)

        sheet_id = order_sheet.id

        # Get a list of truck ids and their earliest departure time
        subq = db.session.query(
            Order.truck_s_number,
            func.min(Order.departure_time).label('mintime')
        ).group_by(Order.truck_s_number).filter(Order.sheet_id == sheet_id) \
            .subquery()

        # Get the first orders for each truck
        first_orders = db.session.query(Order).join(
            subq,
            and_(
                Order.truck_s_number == subq.c.truck_s_number,
                Order.departure_time == subq.c.mintime
            )
        ).all()

        # Create a new sheet file
        book = Workbook()
        sheet = book.active
        now = time.strftime("%d %b %Y %X")
        now_save = time.strftime("%Y-%m-%d-%H-%M-%S")

        # Set the column names
        sheet['A1'] = now
        sheet['C4'] = 'Sno'
        sheet['D4'] = 'Driver Name'
        sheet['E4'] = 'Truck ID'
        sheet['F4'] = 'Terminal'
        sheet['G4'] = 'chassis'
        sheet['H4'] = 'Starting Time'
        sheet['I4'] = 'Delivery Deadline'
        sheet['J4'] = 'Customer'
        sheet['K4'] = 'Container No.'
        sheet['L4'] = 'City'
        sheet['M4'] = 'Container Type'
        sheet['N4'] = 'Shipping company'
        sheet['O4'] = 'Remarks'

        # Styling
        yellow_fill = PatternFill(start_color='FFFF00',
                                  fill_type='solid')

        border = Border(left=Side(border_style='thin', color='000000'),
                        right=Side(border_style='thin', color='000000'),
                        top=Side(border_style='thin', color='000000'),
                        bottom=Side(border_style='thin', color='000000'))

        # Set the styling to each of the column header cells
        for cell in sheet.iter_cols(3, 15, 4, 4):
            cell[0].fill = yellow_fill
            cell[0].font = Font(bold=True)
            cell[0].border = border

        # Set the width of cells that contain long values
        sheet.column_dimensions['D'].width = 20
        sheet.column_dimensions['H'].width = 15
        sheet.column_dimensions['I'].width = 15
        sheet.column_dimensions['N'].width = 20
        sheet.column_dimensions['O'].width = 40

        # Create a row for each truck that has assigned an order
        for count, order in zip(range(5, len(first_orders)+5), first_orders):

            sheet.cell(row=count, column=3).value = \
                order.truck.s_number  # s number
            sheet.cell(row=count, column=4).value = \
                order.truck.others.get('Driver', '')   # driver
            sheet.cell(row=count, column=5).value = \
                order.truck.truck_id  # truck id
            sheet.cell(row=count, column=6).value = \
                order.inl_terminal  # terminal
            sheet.cell(row=count, column=7).value = \
                ''  # chassis?
            sheet.cell(row=count, column=8).value = \
                order.departure_time  # dep time
            sheet.cell(row=count, column=9).value = \
                order.delivery_deadline  # deadline
            sheet.cell(row=count, column=10).value = \
                order.others.get('Client', '')  # client
            sheet.cell(row=count, column=11).value = \
                order.others.get('Container', '')  # container number
            sheet.cell(row=count, column=12).value = \
                order.others.get('City', '')  # city
            sheet.cell(row=count, column=13).value = \
                order.others.get('Unit type', '')  # container type
            sheet.cell(row=count, column=14).value = \
                order.others.get('Ship. comp.', '')  # shipping company
            sheet.cell(row=count, column=15).value = \
                order.truck.others.get('Remarks', '')  # remarks

            # Set the borders of each cell of the row
            for i in range(3, 16):
                sheet.cell(row=count, column=i).border = border

        # Save the file to an io stream
        filename = 'first-rides-' + now_save + '.xlsx'
        file = io.BytesIO()
        book.save(file)
        file.seek(0)

        return send_file(
            file,
            attachment_filename=filename,
            as_attachment=True
            ), 200
